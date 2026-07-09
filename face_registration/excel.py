from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime


class ExcelManager:

    def __init__(self):

        self.file = "database/users.xlsx"

        Path("database").mkdir(
            exist_ok=True
        )


        # Create file if not exist
        if not Path(self.file).exists():

            workbook = Workbook()

            sheet = workbook.active

            sheet.title = "Users"


            sheet.append([
                "Name",
                "Student ID",
                "Major",
                "Role",
                "Image Path",
                "Created At"
            ])


            workbook.save(
                self.file
            )


    def save_user(
        self,
        name,
        student_id,
        major,
        role,
        image_path
    ):

        workbook = load_workbook(
            self.file
        )


        sheet = workbook["Users"]


        sheet.append([
            name,
            student_id,
            major,
            role,
            image_path,
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
        ])


        workbook.save(
            self.file
        )